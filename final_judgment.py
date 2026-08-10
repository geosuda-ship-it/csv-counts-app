"""
Google Colab用：サンプルごとの最終グループ判定ツール

使い方
1. このコード全体をGoogle Colabの1セルに貼り付けて実行する。
2. STEP 2で作成した判定結果ファイルをアップロードする。
   例：20260728_判定結果.xlsx
3. 共通の「グループ集計シート.xlsx」をアップロードする。
   例：グループ集計シート_ver1.xlsx
4. 処理完了後、「日付_最終判定結果.xlsx」と
   「日付_グループ集計結果.xlsx」が自動ダウンロードされる。

入力シート（必須）
・判定結果

必要な列（表記揺れには一部対応）
・Sample（資料No.、資料、試料名でも可）
・判定結果（化学組成グループ、Groupでも可）
・得点（Scoreでも可）
・有効/無効（有効／無効、判定有効性、Validityでも可）

結果集計
・2番目にアップロードしたExcelのレイアウトと書式を維持する。
・D3:F33は、集計値が0の場合は空欄にする。
・G列および34～36行の合計部分は、0も数値として表示する。
・B列の化学組成グループ名には、I列のURLをリンクとして設定する。
"""

from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 利用者設定
# ============================================================

# 個別URLを指定したい場合は追加する。
GROUP_URLS = {
    # "KD-1": "https://example.com/kd-1",
}

# GROUP_URLSにも入力ExcelにもURLがない場合のURL生成規則。
# 不要なら空文字にする。
BASE_URL = "https://sites.google.com/view/obsidian/reference/kyushu/nw-kyushu/"

# 同点併記（KD-1／KD-2）は、テンプレートに各グループの行があるため、
# KD-1とKD-2の双方へ1件ずつ計上する。
COUNT_TIED_RESULT_AS_ONE_CATEGORY = False


# ============================================================
# 共通処理
# ============================================================

def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lstrip("'")


def make_output_file_names(input_name: str) -> tuple[str, str]:
    """判定結果ファイル名から、日付付きの2つの出力名を作る。"""
    input_stem = Path(input_name).stem

    # ブラウザが重複ファイルへ付けた「 (1)」「 (2)」などを除く。
    cleaned_stem = re.sub(
        r"\s*\(\d+\)$",
        "",
        input_stem,
    ).strip()

    date_match = re.match(r"^(\d{8})", cleaned_stem)
    if date_match is None:
        raise ValueError(
            "グループ判定結果ファイル名の先頭から8桁の日付を"
            "読み取れません。\n例：20260728_判定結果.xlsx"
        )

    date_text = date_match.group(1)
    return (
        f"{date_text}_最終判定結果.xlsx",
        f"{date_text}_グループ集計結果.xlsx",
    )


def normalize(value) -> str:
    text = unicodedata.normalize("NFKC", clean_text(value)).lower()
    return re.sub(r"[\s_\-・/／]+", "", text)


def find_column(df: pd.DataFrame, candidates: list[str], required=True):
    actual = {normalize(c): c for c in df.columns}
    for candidate in candidates:
        if normalize(candidate) in actual:
            return actual[normalize(candidate)]
    if required:
        raise KeyError(
            f"必要な列が見つかりません。候補={candidates}、実際の列={list(df.columns)}"
        )
    return None


def is_valid(value) -> bool:
    text = normalize(value)
    if text in {"有効", "valid", "true", "1", "○", "〇"}:
        return True
    if text in {"無効", "invalid", "false", "0", "×", "x", ""}:
        return False
    # 「有効候補」等を許容。ただし「無効」を先に除外する。
    return "有効" in text and "無効" not in text


def confidence(score) -> str:
    if pd.isna(score):
        return ""
    score = float(score)
    if 8 <= score <= 10:
        return "高"
    if 5 <= score < 8:
        return "中"
    if 3 <= score < 5:
        return "低"
    return ""


def fmt_number(value) -> str:
    value = float(value)
    return str(int(value)) if value.is_integer() else f"{value:g}"


def fmt_percent(value) -> str:
    return f"{float(value):.1%}"


def join_values(rows, key, formatter=str) -> str:
    return "／".join(formatter(row[key]) for row in rows)


# ============================================================
# 入力の読込みと整形
# ============================================================

def get_one_excel(uploaded: dict[str, bytes], description: str) -> tuple[str, bytes]:
    excel_files = {
        name: content
        for name, content in uploaded.items()
        if Path(name).suffix.lower() in {".xlsx", ".xlsm"}
    }
    if len(excel_files) != 1:
        raise ValueError(
            f"{description}を1ファイルだけアップロードしてください。"
            f" アップロード済み={list(uploaded)}"
        )
    return next(iter(excel_files.items()))


def create_final_judgment(
    judgment_content: bytes,
    judgment_file_name: str,
    template_content: bytes,
    template_file_name: str,
) -> dict:
    if not judgment_content:
        raise ValueError("STEP 2で作成した判定結果ファイルが空です。")
    if not template_content:
        raise ValueError("グループ集計シート.xlsxが空です。")

    input_name = judgment_file_name
    input_bytes = judgment_content
    JUDGMENT_OUTPUT_FILENAME, SUMMARY_OUTPUT_FILENAME = make_output_file_names(
        input_name
    )

    template_name = template_file_name
    template_bytes = template_content
    
    book = pd.ExcelFile(io.BytesIO(input_bytes))
    if "判定結果" not in book.sheet_names:
        raise KeyError(
            f"『判定結果』シートがありません。実際のシート={book.sheet_names}"
        )
    
    raw = pd.read_excel(book, sheet_name="判定結果")
    
    sample_col = find_column(raw, ["Sample", "資料No.", "資料No", "資料", "試料名"])
    analysis_col = find_column(
        raw, ["分析値No.", "分析値No", "分析No.", "分析No", "測定No."], required=False
    )
    group_col = find_column(raw, ["判定結果", "化学組成グループ", "Group"])
    score_col = find_column(raw, ["得点", "Score"])
    valid_col = find_column(raw, ["有効/無効", "有効／無効", "判定有効性", "Validity"])
    
    data = pd.DataFrame({
        "資料No.": raw[sample_col].map(clean_text),
        "分析値No.": raw[analysis_col] if analysis_col else range(1, len(raw) + 1),
        "判定結果": raw[group_col].map(clean_text),
        "得点": pd.to_numeric(raw[score_col], errors="coerce"),
        "有効/無効": raw[valid_col].map(clean_text),
    })
    data = data[data["資料No."] != ""].reset_index(drop=True)
    data["有効"] = data["有効/無効"].map(is_valid)
    
    # 有効行はグループ名と3～10点の得点を必要とする。
    bad_valid = data[
        data["有効"]
        & ((data["判定結果"] == "") | ~data["得点"].between(3, 10, inclusive="both"))
    ]
    if not bad_valid.empty:
        raise ValueError(
            "有効行に空の判定結果、または3～10点以外の得点があります。\n"
            + bad_valid.to_string(index=False)
        )
    
    
    def extract_urls(excel: pd.ExcelFile) -> dict[str, str]:
        result = dict(GROUP_URLS)
        for sheet in excel.sheet_names:
            try:
                df = pd.read_excel(excel, sheet_name=sheet)
                gcol = find_column(df, ["化学組成グループ", "判定結果", "Group"], required=False)
                ucol = find_column(df, ["URL", "リンク", "Link"], required=False)
                if gcol and ucol:
                    for group, url in zip(df[gcol], df[ucol]):
                        group, url = clean_text(group), clean_text(url)
                        if group and url.startswith(("http://", "https://")):
                            result[group] = url
            except Exception:
                # URL表でないシートは読み飛ばす。
                pass
        return result
    
    
    url_map = extract_urls(book)
    
    
    def group_url(group: str) -> str:
        if group in url_map:
            return url_map[group]
        if BASE_URL and "／" not in group:
            return BASE_URL.rstrip("/") + "/" + group.lower()
        return ""
    
    
    # ============================================================
    # 集計と最終判定
    # ============================================================
    
    aggregate_rows = []
    final_rows = []
    
    for sample, sample_df in data.groupby("資料No.", sort=False):
        valid_df = sample_df[sample_df["有効"]].copy()
        valid_count = len(valid_df)
    
        if valid_count == 0:
            final_rows.append({
                "資料No.": sample, "最終判定": "判定不能", "判定確度": "",
                "最大得点": "", "判定回数": "", "信頼性": "",
                "次候補": "", "次候補_最大得点": "",
                "次候補_判定回数": "", "次候補_信頼性": "",
            })
            continue
    
        per_group = []
        for group, group_df in valid_df.groupby("判定結果", sort=False):
            score_sum = float(group_df["得点"].sum())
            row = {
                "資料No.": sample,
                "化学組成グループ": group,
                "有効判定数": valid_count,
                "判定回数": int(len(group_df)),
                "得点合計": score_sum,
                "信頼性": score_sum / (valid_count * 10),
                "最大得点": float(group_df["得点"].max()),
                "8点以上の有無": "有" if (group_df["得点"] >= 8).any() else "無",
            }
            per_group.append(row)
    
        high_score_mode = bool((valid_df["得点"] >= 8).any())
        if high_score_mode:
            rank_key = lambda r: (r["最大得点"], r["信頼性"])
            rule_name = "最大得点→信頼性"
        else:
            rank_key = lambda r: (r["判定回数"], r["信頼性"], r["最大得点"])
            rule_name = "判定回数→信頼性→最大得点"
    
        ranked = sorted(per_group, key=rank_key, reverse=True)
        for priority, row in enumerate(ranked, 1):
            row["優先順位"] = priority
            row["判定規則"] = rule_name
            aggregate_rows.append(row)
    
        best_key = rank_key(ranked[0])
        winners = [r for r in ranked if rank_key(r) == best_key]
    
        if len(winners) > 1:
            final_name = join_values(winners, "化学組成グループ")
            next_rows = []
        else:
            final_name = winners[0]["化学組成グループ"]
            remaining = [r for r in ranked if r not in winners]
            if remaining:
                second_key = rank_key(remaining[0])
                next_rows = [r for r in remaining if rank_key(r) == second_key]
            else:
                next_rows = []
    
        final_rows.append({
            "資料No.": sample,
            "最終判定": final_name,
            "判定確度": join_values(winners, "最大得点", confidence),
            "最大得点": join_values(winners, "最大得点", fmt_number),
            "判定回数": join_values(winners, "判定回数", lambda x: str(int(x))),
            "信頼性": join_values(winners, "信頼性", fmt_percent),
            "次候補": join_values(next_rows, "化学組成グループ") if next_rows else "",
            "次候補_最大得点": join_values(next_rows, "最大得点", fmt_number) if next_rows else "",
            "次候補_判定回数": join_values(next_rows, "判定回数", lambda x: str(int(x))) if next_rows else "",
            "次候補_信頼性": join_values(next_rows, "信頼性", fmt_percent) if next_rows else "",
        })
    
    aggregate_df = pd.DataFrame(aggregate_rows)
    final_df = pd.DataFrame(final_rows)
    
    # 結果集計用データ
    summary_source = []
    for row in final_rows:
        result = row["最終判定"]
        if result == "判定不能":
            summary_source.append((result, ""))
        elif COUNT_TIED_RESULT_AS_ONE_CATEGORY or "／" not in result:
            # 併記の場合の確度は全候補で同じ（順位キーが完全一致）である。
            summary_source.append((result, row["判定確度"].split("／")[0]))
        else:
            groups = result.split("／")
            levels = row["判定確度"].split("／")
            summary_source.extend(zip(groups, levels))
    
    summary_counts = {}
    for group, level in summary_source:
        summary_counts.setdefault(group, {"高": 0, "中": 0, "低": 0})
        if level in {"高", "中", "低"}:
            summary_counts[group][level] += 1
    
    
    # ============================================================
    # Excel出力（計算値はすべて固定値）
    # ============================================================
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    
    def add_sheet(name: str, df: pd.DataFrame, hyperlink_column=None):
        ws = wb.create_sheet(name)
        if df.empty and len(df.columns) == 0:
            ws["A1"] = "データなし"
            return ws
        for col_no, column in enumerate(df.columns, 1):
            cell = ws.cell(1, col_no, column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_no, row in enumerate(df.itertuples(index=False, name=None), 2):
            for col_no, value in enumerate(row, 1):
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row_no, col_no, value)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            if hyperlink_column and hyperlink_column in df.columns:
                col_no = df.columns.get_loc(hyperlink_column) + 1
                cell = ws.cell(row_no, col_no)
                url = group_url(clean_text(cell.value))
                if url:
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
                    cell.border = border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_no, column in enumerate(df.columns, 1):
            values = [clean_text(column)] + [clean_text(v) for v in df.iloc[:, col_no - 1].head(500)]
            width = min(max(max(map(len, values)) + 2, 10), 35)
            ws.column_dimensions[get_column_letter(col_no)].width = width
        return ws
    
    
    source_out = data.drop(columns=["有効"])
    # 「判定結果」を最初に作成し、Excelで一番左のシートにする。
    result_ws = add_sheet("判定結果", final_df, "最終判定")
    add_sheet("元データ", source_out)
    agg_ws = add_sheet("グループ別集計", aggregate_df, "化学組成グループ")
    
    # 判定結果シートは、見出し・文字・数字をすべて中央揃えにする。
    for row in result_ws.iter_rows(
        min_row=1,
        max_row=result_ws.max_row,
        min_col=1,
        max_col=result_ws.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
    
    # 数値表示
    if not aggregate_df.empty:
        reliability_col = aggregate_df.columns.get_loc("信頼性") + 1
        for row in range(2, agg_ws.max_row + 1):
            agg_ws.cell(row, reliability_col).number_format = "0.0%"
    
    # 判定結果は可読性を優先して二段見出しに近い配色を付ける。
    for col in range(1, result_ws.max_column + 1):
        if col >= 7:
            result_ws.cell(1, col).fill = PatternFill("solid", fgColor="548235")
    
    judgment_output = io.BytesIO()
    wb.save(judgment_output)
    
    
    # ============================================================
    # 結果集計テンプレートへの固定値書込み
    # ============================================================
    
    summary_wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
    summary_ws = summary_wb.active
    
    if summary_ws.max_row < 36 or summary_ws.max_column < 10:
        raise ValueError(
            "グループ集計シートのレイアウトが想定と異なります。"
            " 少なくともA1:J36の範囲が必要です。"
        )
    
    # J3:J32にグループ名、B33に「判定不能」があるテンプレートを想定する。
    template_rows = {}
    for row_no in range(3, 34):
        group = clean_text(summary_ws.cell(row_no, 10).value)
        if not group:
            b_value = clean_text(summary_ws.cell(row_no, 2).value)
            if not b_value.startswith("="):
                group = b_value
        if group:
            template_rows[group] = row_no
    
    if "判定不能" not in template_rows:
        template_rows["判定不能"] = 33
    
    # テンプレートに存在しない判定結果を明示して、集計漏れを防止する。
    missing_groups = sorted(
        group for group in summary_counts
        if group not in template_rows and group != "判定不能"
    )
    if missing_groups:
        raise ValueError(
            "グループ集計シートに次の化学組成グループがありません："
            + "、".join(missing_groups)
        )
    
    for group, row_no in template_rows.items():
        counts = summary_counts.get(group, {"高": 0, "中": 0, "低": 0})
    
        # D3:F33は0を空欄にする。
        for col_no, level in zip(range(4, 7), ["高", "中", "低"]):
            value = int(counts[level])
            summary_ws.cell(row_no, col_no).value = value if value != 0 else None
    
        if group == "判定不能":
            total = sum(1 for result, _ in summary_source if result == "判定不能")
        else:
            total = int(sum(counts.values()))
    
        # 合計列は0も表示する。
        summary_ws.cell(row_no, 7).value = total
    
        # B列は表示名を固定値にし、I列のURLを実体リンクとして設定する。
        if row_no <= 32:
            url = clean_text(summary_ws.cell(row_no, 9).value)
            display_name = clean_text(summary_ws.cell(row_no, 10).value) or group
            summary_ws.cell(row_no, 2).value = display_name
            if url.startswith(("http://", "https://")):
                summary_ws.cell(row_no, 2).hyperlink = url
    
    # 下部の合計欄は、0を含めて固定値で書き込む。
    high_total = sum(values["高"] for values in summary_counts.values())
    middle_total = sum(values["中"] for values in summary_counts.values())
    low_total = sum(values["低"] for values in summary_counts.values())
    judged_total = high_total + middle_total + low_total
    unclassifiable_total = sum(1 for result, _ in summary_source if result == "判定不能")
    sample_total = len(final_rows)
    judgment_rate = judged_total / sample_total if sample_total else 0
    
    summary_ws["D34"] = int(high_total)
    summary_ws["E34"] = int(middle_total)
    summary_ws["F34"] = int(low_total)
    summary_ws["G34"] = int(judged_total)
    summary_ws["G35"] = int(sample_total)
    summary_ws["G36"] = float(judgment_rate)
    summary_ws["G36"].number_format = "0.0%"
    
    summary_output = io.BytesIO()
    summary_wb.save(summary_output)

    return {
        "judgment_file_name": JUDGMENT_OUTPUT_FILENAME,
        "judgment_content": judgment_output.getvalue(),
        "summary_file_name": SUMMARY_OUTPUT_FILENAME,
        "summary_content": summary_output.getvalue(),
        "sample_count": int(sample_total),
        "judged_count": int(judged_total),
        "unclassifiable_count": int(unclassifiable_total),
        "judgment_rate": float(judgment_rate),
    }
