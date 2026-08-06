"""
Google Colab用：EDX CSVから定量値Excelまでを一括作成する。

処理:
1. CSVから強度測定値を抽出
2. 同日のQC-2/QC2（obart_prec3_air）からドリフト補正係数を算出
3. K～Feはドリフト補正、Zn～NbはさらにAg内標準補正
4. 検量線定数B・Cから定量値を算出
5. Y、Zr、Nbに重なり補正を適用

Colabでは、このコード全体を1つのセルに貼り付けて実行する。
CSVをアップロードすると、最終Excelが自動的にダウンロードされる。
"""

from __future__ import annotations

import csv
import io
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ELEMENTS = ("K", "Ca", "Mn", "Fe", "Zn", "Rb", "Sr", "Y", "Zr", "Nb", "Ag")
QUANT_ELEMENTS = ELEMENTS[:-1]
DIRECT_ELEMENTS = {"K", "Ca", "Mn", "Fe"}
AG_NORMALIZED_ELEMENTS = {"Zn", "Rb", "Sr", "Y", "Zr", "Nb"}
QC_MODE = "obart_prec3_air"

REFERENCE_INTENSITY = {
    "K": 4.27826,
    "Ca": 0.15422,
    "Mn": 0.90407,
    "Fe": 3.15459,
    "Zn": 0.01778,
    "Rb": 0.22959,
    "Sr": 0.07396,
    "Y": 0.09792,
    "Zr": 0.17312,
    "Nb": 0.10719,
    "Ag": 1.53429,
}

CALIBRATION_B = {
    "K": 9312.59,
    "Ca": 30145.8,
    "Mn": 746.088,
    "Fe": 4061.51,
    "Zn": 7048.49,
    "Rb": 1360.26,
    "Sr": 1112.8,
    "Y": 898.725,
    "Zr": 810.974,
    "Nb": 737.066,
}

CALIBRATION_C = {
    "K": -1378.82,
    "Ca": -204.571,
    "Mn": -315.439,
    "Fe": -5513.26,
    "Zn": -44.4817,
    "Rb": -16.0436,
    "Sr": -11.9096,
    "Y": -12.6423,
    "Zr": -18.6844,
    "Nb": -31.1933,
}

# 対象元素: (重なり元素, 重なり補正係数)
OVERLAP_CORRECTION = {
    "Y": ("Rb", -0.113036),
    "Zr": ("Sr", -0.121299),
    "Nb": ("Y", -0.161034),
}


def decode_csv(csv_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return csv_bytes.decode(encoding)
        except UnicodeDecodeError:
            pass
    raise ValueError("CSVの文字コードを判定できませんでした。")


def parse_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = str(value or "").strip()
    for date_format in (
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    raise ValueError(f"測定日時を読めません: {value!r}")


def normalize_qc_name(value) -> str:
    text = str(value or "").strip().upper()
    for character in ("-", "－", "−", "‐", "‑", "–", "—", " ", "　"):
        text = text.replace(character, "")
    return text


def is_qc(measurement: dict) -> bool:
    return (
        normalize_qc_name(measurement["sample"]) == "QC2"
        and measurement["mode"].strip().lower() == QC_MODE.lower()
    )


def extract_measurements(csv_bytes: bytes) -> list[dict]:
    rows = list(csv.reader(io.StringIO(decode_csv(csv_bytes))))
    measurements = []
    in_intensity_block = False

    for row_number, source_row in enumerate(rows, start=1):
        row = source_row + [""] * max(0, 14 - len(source_row))

        # μが文字化けしても「cps/」で強度ブロックを識別する。
        if sum("cps/" in cell.strip().lower() for cell in row) >= 5:
            in_intensity_block = True
            continue

        if not any(cell.strip() for cell in row):
            in_intensity_block = False
            continue

        if not in_intensity_block:
            continue

        if not all(row[column].strip() for column in range(3)):
            continue

        intensities = {}
        for column, element in enumerate(ELEMENTS, start=3):
            text = row[column].strip()
            if not text:
                raise ValueError(
                    f"CSV {row_number}行目の{element}強度が空欄です。"
                )
            try:
                intensities[element] = float(text)
            except ValueError as exc:
                raise ValueError(
                    f"CSV {row_number}行目の{element}強度を数値として読めません。"
                ) from exc

        measurements.append(
            {
                "sample": row[0].strip(),
                "mode": row[1].strip(),
                "measured_at": parse_datetime(row[2]),
                "intensities": intensities,
            }
        )

    if not measurements:
        raise ValueError("CSV内に強度測定データが見つかりません。")
    return measurements


def calculate_quantitative_values(corrected: dict[str, float]) -> dict[str, float]:
    values = {}

    for element in ("K", "Ca", "Mn", "Fe", "Zn", "Rb", "Sr"):
        values[element] = (
            CALIBRATION_B[element] * corrected[element]
            + CALIBRATION_C[element]
        )

    overlap_element, coefficient = OVERLAP_CORRECTION["Y"]
    values["Y"] = (
        CALIBRATION_B["Y"] * corrected["Y"]
        + CALIBRATION_C["Y"]
        + values[overlap_element] * coefficient
    )

    overlap_element, coefficient = OVERLAP_CORRECTION["Zr"]
    values["Zr"] = (
        CALIBRATION_B["Zr"] * corrected["Zr"]
        + CALIBRATION_C["Zr"]
        + values[overlap_element] * coefficient
    )

    overlap_element, coefficient = OVERLAP_CORRECTION["Nb"]
    values["Nb"] = (
        CALIBRATION_B["Nb"] * corrected["Nb"]
        + CALIBRATION_C["Nb"]
        + values[overlap_element] * coefficient
    )
    return values


def calculate_all(measurements: list[dict]):
    qc_by_date = defaultdict(list)
    for measurement in measurements:
        if is_qc(measurement):
            qc_by_date[measurement["measured_at"].date()].append(measurement)

    for qc_list in qc_by_date.values():
        qc_list.sort(key=lambda item: item["measured_at"])

    if not qc_by_date:
        raise ValueError(
            "QC-2またはQC2かつobart_prec3_airのQC測定が見つかりません。"
        )

    coefficient_rows = []
    coefficients_by_date = defaultdict(list)

    for measurement_date in sorted(qc_by_date):
        for qc_number, qc in enumerate(qc_by_date[measurement_date], start=1):
            coefficients = {}
            for element in ELEMENTS:
                qc_intensity = qc["intensities"][element]
                if qc_intensity == 0:
                    raise ValueError(
                        f"{qc['measured_at']}の{element} QC強度が0です。"
                    )
                coefficients[element] = (
                    REFERENCE_INTENSITY[element] / qc_intensity
                )

            coefficients_by_date[measurement_date].append(
                (qc_number, qc, coefficients)
            )
            coefficient_rows.append(
                {
                    "date": measurement_date,
                    "qc_number": qc_number,
                    "qc": qc,
                    "coefficients": coefficients,
                }
            )

    corrected_rows = []
    missing_qc_dates = set()

    for sample in measurements:
        if is_qc(sample):
            continue

        measurement_date = sample["measured_at"].date()
        daily_coefficients = coefficients_by_date.get(measurement_date, [])
        if not daily_coefficients:
            missing_qc_dates.add(measurement_date)
            continue

        for qc_number, qc, coefficients in daily_coefficients:
            drift_corrected = {
                element: sample["intensities"][element] * coefficients[element]
                for element in ELEMENTS
            }

            corrected_ag = drift_corrected["Ag"]
            if corrected_ag == 0:
                raise ValueError(
                    f"{sample['sample']} / QC{qc_number}: "
                    "ドリフト補正後Ag強度が0です。"
                )

            corrected = {}
            for element in ELEMENTS:
                if element in DIRECT_ELEMENTS or element == "Ag":
                    corrected[element] = drift_corrected[element]
                elif element in AG_NORMALIZED_ELEMENTS:
                    corrected[element] = drift_corrected[element] / corrected_ag

            corrected_rows.append(
                {
                    "sample": sample,
                    "date": measurement_date,
                    "qc_number": qc_number,
                    "qc": qc,
                    "corrected": corrected,
                    "quantitative": calculate_quantitative_values(corrected),
                }
            )

    if not corrected_rows:
        raise ValueError("同日のQCを適用できる試料がありません。")

    return (
        coefficient_rows,
        corrected_rows,
        sorted(missing_qc_dates),
    )


def style_sheet(sheet, datetime_columns=(), date_columns=()):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="游ゴシック", size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    for column in datetime_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd h:mm"
    for column in date_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 11), 22
        )


def create_workbook(coefficient_rows, corrected_rows, missing_dates) -> Workbook:
    workbook = Workbook()

    qc_sheet = workbook.active
    qc_sheet.title = "QC補正係数"
    qc_sheet.append(
        [
            "測定日",
            "QC番号",
            "QC sample",
            "QC測定モード",
            "QC測定日時",
            *[f"{element}_補正係数" for element in ELEMENTS],
        ]
    )
    for row in coefficient_rows:
        qc_sheet.append(
            [
                row["date"],
                row["qc_number"],
                row["qc"]["sample"],
                row["qc"]["mode"],
                row["qc"]["measured_at"],
                *[row["coefficients"][element] for element in ELEMENTS],
            ]
        )
    style_sheet(qc_sheet, datetime_columns=("E",), date_columns=("A",))
    for row in qc_sheet.iter_rows(min_row=2, min_col=6, max_col=16):
        for cell in row:
            cell.number_format = "0.000000"

    corrected_sheet = workbook.create_sheet("補正後強度")
    corrected_sheet.append(
        [
            "sample",
            "測定モード",
            "試料測定日時",
            "測定日",
            "QC番号",
            "QC測定日時",
            "K_ドリフト補正後",
            "Ca_ドリフト補正後",
            "Mn_ドリフト補正後",
            "Fe_ドリフト補正後",
            "Zn_Ag内標準後",
            "Rb_Ag内標準後",
            "Sr_Ag内標準後",
            "Y_Ag内標準後",
            "Zr_Ag内標準後",
            "Nb_Ag内標準後",
            "Ag_ドリフト補正後",
        ]
    )
    for row in corrected_rows:
        corrected_sheet.append(
            [
                row["sample"]["sample"],
                row["sample"]["mode"],
                row["sample"]["measured_at"],
                row["date"],
                row["qc_number"],
                row["qc"]["measured_at"],
                *[row["corrected"][element] for element in ELEMENTS],
            ]
        )
    style_sheet(
        corrected_sheet,
        datetime_columns=("C", "F"),
        date_columns=("D",),
    )
    for row in corrected_sheet.iter_rows(min_row=2, min_col=7, max_col=17):
        for cell in row:
            cell.number_format = "0.000000"

    quantitative_sheet = workbook.create_sheet("定量値")
    quantitative_sheet.append(
        [
            "sample",
            "測定モード",
            "試料測定日時",
            *[f"{element}_ppm" for element in QUANT_ELEMENTS],
        ]
    )
    for row in corrected_rows:
        quantitative_sheet.append(
            [
                row["sample"]["sample"],
                row["sample"]["mode"],
                row["sample"]["measured_at"],
                *[row["quantitative"][element] for element in QUANT_ELEMENTS],
            ]
        )
    style_sheet(
        quantitative_sheet,
        datetime_columns=("C",),
    )
    for row in quantitative_sheet.iter_rows(min_row=2, min_col=4, max_col=13):
        for cell in row:
            cell.number_format = "0.000"

    constants_sheet = workbook.create_sheet("計算定数")
    constants_sheet.append(
        [
            "元素",
            "検量線定数B",
            "検量線定数C",
            "重なり元素",
            "重なり補正係数",
        ]
    )
    for element in QUANT_ELEMENTS:
        overlap_element, overlap_coefficient = OVERLAP_CORRECTION.get(
            element, ("", "")
        )
        constants_sheet.append(
            [
                element,
                CALIBRATION_B[element],
                CALIBRATION_C[element],
                overlap_element,
                overlap_coefficient,
            ]
        )
    style_sheet(constants_sheet)
    for row in constants_sheet.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"

    if missing_dates:
        warning_sheet = workbook.create_sheet("警告")
        warning_sheet.append(["同日のQCがなく、出力から除外した測定日"])
        for missing_date in missing_dates:
            warning_sheet.append([missing_date])
        style_sheet(warning_sheet, date_columns=("A",))

    # 「定量値」シートを一番左へ移動し、Excelを開いたときにも表示する。
    workbook.move_sheet(
        quantitative_sheet,
        offset=-workbook.index(quantitative_sheet),
    )
    workbook.active = workbook.index(quantitative_sheet)

    return workbook


def convert_csv(csv_bytes: bytes) -> bytes:
    """将来Webアプリでも再利用できる、bytes入出力の一括変換関数。"""
    measurements = extract_measurements(csv_bytes)
    coefficient_rows, corrected_rows, missing_dates = calculate_all(measurements)
    workbook = create_workbook(
        coefficient_rows,
        corrected_rows,
        missing_dates,
    )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def run_colab():
    from google.colab import files

    uploaded = files.upload()
    csv_names = [
        name for name in uploaded if name.lower().endswith(".csv")
    ]
    if not csv_names:
        raise ValueError(".csvファイルをアップロードしてください。")

    input_name = csv_names[0]
    output_name = f"{Path(input_name).stem}_補正後強度_定量値.xlsx"
    output_bytes = convert_csv(uploaded[input_name])

    with open(output_name, "wb") as handle:
        handle.write(output_bytes)

    print(f"完了: {output_name}")
    files.download(output_name)


def run_local(input_path: str, output_path: str | None = None):
    input_file = Path(input_path)
    output_file = (
        Path(output_path)
        if output_path
        else input_file.with_name(
            f"{input_file.stem}_補正後強度_定量値.xlsx"
        )
    )
    output_file.write_bytes(convert_csv(input_file.read_bytes()))
    print(f"完了: {output_file}")


def running_in_colab() -> bool:
    try:
        import google.colab  # noqa: F401
        return True
    except ImportError:
        return False


if __name__ == "__main__":
    if running_in_colab():
        run_colab()
    elif len(sys.argv) >= 2:
        run_local(sys.argv[1], sys.argv[2] if len(sys.argv) >= 3 else None)
    else:
        print("このコードをGoogle Colabで実行してください。")
