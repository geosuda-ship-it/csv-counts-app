"""
Google Colab用
各定量値の化学組成グループ判定ツール

使い方
------
1. コード全体をGoogle Colabの1セルに貼り付けて実行
2. STEP 1で作成した定量値ファイルをアップロード
   例：20260728_定量値.xlsx
3. 続いて，共通の判別楕円パラメーターファイルをアップロード
   例：長崎大学_判別楕円パラメータ_ver1.xlsx
4. グループ判定結果ファイル
   （日付_判定結果.xlsx）が自動的にダウンロードされる

判定方針
--------
・「定量値」シートの1行を1組の定量分析値として個別に判定する
・Sample名が同じでも，各行を別々に判定する
・各定量分析値について，次の5種の判定用パラメータを算出する
  ln(100×K/Ca)，Rb*，Sr*，Y*，Zr*
・D1～D3のすべてにおいて，同じ化学組成グループの
  判別楕円内に入った場合に「有効」とする
・D1～D3のいずれかで楕円外となった場合は「無効」とする
・有効な候補について，D1～D10のうち楕円内となった数を
  「得点」とする
・有効候補の中で得点が最大のグループを最終判定とする
・同点の場合は，パラメータ表で先に並ぶグループを採用する

Diagram
-------
D1  ：Rb* vs Sr*
D2  ：Rb* vs ln(100×K/Ca)
D3  ：Sr* vs ln(100×K/Ca)
D4  ：Rb* vs Y*
D5  ：Rb* vs Zr*
D6  ：Sr* vs Y*
D7  ：Sr* vs Zr*
D8  ：Y* vs Zr*
D9  ：Y* vs ln(100×K/Ca)
D10 ：Zr* vs ln(100×K/Ca)

判別楕円パラメータの必須列
--------------------------
Diagram, Group, X, Y, x0, y0, a, b, theta,
Mandatory または Required

Mandatory／Required列は，
D1～D3を1，D4～D10を0とする。
"""

from __future__ import annotations

import io
import math
import re
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 1．基本設定
# ============================================================

INPUT_SHEET_NAME = "定量値"

ELLIPSE_LIMIT = 1.0

INSIDE_MARK = "○"
OUTSIDE_MARK = "×"


def make_output_file_name(input_name: str) -> str:
    """定量値ファイル名から「日付_判定結果.xlsx」を作る。"""
    input_stem = Path(input_name).stem

    # ブラウザが重複ファイルへ付けた「 (1)」「 (2)」などを除く。
    cleaned_stem = re.sub(
        r"\s*\(\d+\)$",
        "",
        input_stem,
    ).strip()

    date_match = re.match(
        r"^(\d{8})",
        cleaned_stem,
    )
    if date_match is None:
        raise ValueError(
            "定量値ファイル名の先頭から8桁の日付を読み取れません。\n"
            "例：20260511_定量値.xlsx"
        )

    return f"{date_match.group(1)}_判定結果.xlsx"


# ============================================================
# 2．D1～D10の正式仕様
# ============================================================

EXPECTED_DIAGRAM_VARIABLES = {
    1: ("Rb*", "Sr*"),
    2: ("Rb*", "ln(100*K/Ca)"),
    3: ("Sr*", "ln(100*K/Ca)"),
    4: ("Rb*", "Y*"),
    5: ("Rb*", "Zr*"),
    6: ("Sr*", "Y*"),
    7: ("Sr*", "Zr*"),
    8: ("Y*", "Zr*"),
    9: ("Y*", "ln(100*K/Ca)"),
    10: ("Zr*", "ln(100*K/Ca)"),
}

EXPECTED_REQUIRED_DIAGRAMS = {1, 2, 3}


# ============================================================
# 3．共通関数
# ============================================================

def clean_text(value) -> str:
    """
    セル値を文字列に変換し，
    前後の空白と先頭のアポストロフィを除く。
    """
    if pd.isna(value):
        return ""

    return str(value).strip().lstrip("'")


def normalize_column_name(value) -> str:
    """
    列名の比較用に，
    半角空白，全角空白，改行を除いて小文字化する。
    """
    return (
        str(value)
        .strip()
        .replace(" ", "")
        .replace("　", "")
        .replace("\n", "")
        .replace("\r", "")
        .lower()
    )


def normalize_parameter_name(value) -> str:
    """
    パラメータ名の表記ゆれを統一する。

    例
    --
    ln（100×K/Ca）
    ln(100*K/Ca)
    ln(100×K/Ca)

    をすべて ln(100*K/Ca) として扱う。
    """
    text = clean_text(value)

    text = (
        text
        .replace("（", "(")
        .replace("）", ")")
        .replace("×", "*")
        .replace("＊", "*")
        .replace("＋", "+")
        .replace("−", "-")
        .replace("－", "-")
        .replace(" ", "")
        .replace("　", "")
    )

    normalized_lower = text.lower()

    if normalized_lower == "rb*":
        return "Rb*"

    if normalized_lower == "sr*":
        return "Sr*"

    if normalized_lower == "y*":
        return "Y*"

    if normalized_lower == "zr*":
        return "Zr*"

    if normalized_lower in {
        "ln(100*k/ca)",
        "ln(100k/ca)",
    }:
        return "ln(100*K/Ca)"

    return text


def get_single_uploaded_excel(
    uploaded: dict[str, bytes],
    description: str,
) -> tuple[str, bytes]:
    """
    アップロードされたExcelファイルが
    1ファイルだけか確認する。
    """

    excel_files = {
        name: content
        for name, content in uploaded.items()
        if Path(name).suffix.lower() in {
            ".xlsx",
            ".xlsm",
        }
    }

    if len(excel_files) != 1:
        raise ValueError(
            f"{description}を1ファイルだけアップロードしてください。\n"
            f"アップロードされたファイル：{list(uploaded.keys())}"
        )

    name, content = next(iter(excel_files.items()))

    return name, content


def find_column(
    df: pd.DataFrame,
    candidates: list[str],
) -> str:
    """
    候補名から，実際の列名を探す。
    """

    normalized_columns = {
        normalize_column_name(column): column
        for column in df.columns
    }

    for candidate in candidates:
        key = normalize_column_name(candidate)

        if key in normalized_columns:
            return normalized_columns[key]

    raise KeyError(
        "必要な列が見つかりません。\n"
        f"候補：{candidates}\n"
        f"実際の列：{list(df.columns)}"
    )


def detect_parameter_columns(
    parameter_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    判別楕円パラメータ表の列名を統一する。

    MandatoryとRequiredは，
    どちらもPython内部ではRequiredとして扱う。
    """

    rename_map = {}

    for column in parameter_df.columns:
        original = str(column).strip()
        normalized = normalize_column_name(column)

        if normalized.startswith("diagram"):
            rename_map[column] = "Diagram"

        elif normalized in {
            "group",
            "グループ",
            "化学組成グループ",
        }:
            rename_map[column] = "Group"

        elif original == "X" or normalized == "x":
            rename_map[column] = "X"

        elif original == "Y" or normalized == "y":
            rename_map[column] = "Y"

        elif normalized in {
            "x0",
            "x₀",
        }:
            rename_map[column] = "x0"

        elif normalized in {
            "y0",
            "y₀",
        }:
            rename_map[column] = "y0"

        elif normalized == "a":
            rename_map[column] = "a"

        elif normalized == "b":
            rename_map[column] = "b"

        elif normalized in {
            "theta",
            "θ",
            "角度",
            "回転角",
        }:
            rename_map[column] = "theta"

        elif (
            normalized.startswith("required")
            or normalized.startswith("mandatory")
            or normalized in {
                "必須",
                "必須図",
                "必須散布図",
            }
        ):
            rename_map[column] = "Required"

    renamed_df = parameter_df.rename(
        columns=rename_map
    )

    # RequiredとMandatoryの両方が存在し，
    # 両方ともRequiredへ変換された場合の重複を検出する。
    duplicated_columns = (
        renamed_df.columns[
            renamed_df.columns.duplicated()
        ]
        .tolist()
    )

    if duplicated_columns:
        raise ValueError(
            "判別楕円パラメータの列名が重複しています。\n"
            "RequiredとMandatoryの両方がある場合は，"
            "どちらか一方を削除してください。\n"
            f"重複列：{duplicated_columns}"
        )

    return renamed_df


def parse_required(value) -> int:
    """
    Mandatory／Required列を0または1に変換する。

    次の値を1として扱う。
    1，TRUE，必須，○，〇，yes，y
    """

    if pd.isna(value):
        return 0

    if isinstance(value, bool):
        return int(value)

    text = clean_text(value).lower()

    if text in {
        "1",
        "true",
        "必須",
        "○",
        "〇",
        "yes",
        "y",
    }:
        return 1

    if text in {
        "0",
        "false",
        "不要",
        "×",
        "no",
        "n",
        "",
    }:
        return 0

    try:
        numeric_value = float(value)

    except (TypeError, ValueError) as error:
        raise ValueError(
            "Mandatory／Required列に，"
            f"0または1として解釈できない値があります：{value}"
        ) from error

    return 1 if numeric_value != 0 else 0


# ============================================================
# 4．判別楕円パラメータの整形・検査
# ============================================================

def prepare_parameters(
    raw_parameter_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    判別楕円パラメータを整形・検査する。

    各化学組成グループについて，
    ・D1～D10が1行ずつ存在する
    ・X・Yの組合せが所定の仕様どおりである
    ・D1～D3だけがMandatory／Required=1である
    ことを確認する。
    """

    df = detect_parameter_columns(
        raw_parameter_df
    ).copy()

    required_columns = [
        "Diagram",
        "Group",
        "X",
        "Y",
        "x0",
        "y0",
        "a",
        "b",
        "theta",
        "Required",
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            "判別楕円パラメータに必要な列がありません。\n"
            f"不足列：{missing_columns}\n"
            f"実際の列：{list(df.columns)}\n\n"
            "MandatoryまたはRequiredという列が必要です。"
        )

    df = df[required_columns].copy()

    # DiagramまたはGroupが空欄の行は除外する。
    df = df.dropna(
        subset=[
            "Diagram",
            "Group",
        ]
    )

    df["Group"] = df["Group"].map(
        clean_text
    )

    df["X"] = df["X"].map(
        normalize_parameter_name
    )

    df["Y"] = df["Y"].map(
        normalize_parameter_name
    )

    # Groupが実質的に空欄の行を除外する。
    df = df[
        df["Group"] != ""
    ].copy()

    numeric_columns = [
        "Diagram",
        "x0",
        "y0",
        "a",
        "b",
        "theta",
    ]

    for column in numeric_columns:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    df["Required"] = df["Required"].map(
        parse_required
    )

    invalid_rows = df[
        df[numeric_columns].isna().any(axis=1)
    ]

    if not invalid_rows.empty:
        raise ValueError(
            "判別楕円パラメータに，"
            "数値として読めないセルがあります。\n\n"
            f"{invalid_rows.to_string(index=False)}"
        )

    df["Diagram"] = df["Diagram"].astype(int)
    df["Required"] = df["Required"].astype(int)

    invalid_diagram_rows = df[
        ~df["Diagram"].between(1, 10)
    ]

    if not invalid_diagram_rows.empty:
        raise ValueError(
            "Diagram列には1～10を入力してください。\n\n"
            f"{invalid_diagram_rows.to_string(index=False)}"
        )

    if (df["a"] <= 0).any():
        invalid_a_rows = df[df["a"] <= 0]

        raise ValueError(
            "判別楕円パラメータのaには，"
            "0より大きい値が必要です。\n\n"
            f"{invalid_a_rows.to_string(index=False)}"
        )

    if (df["b"] <= 0).any():
        invalid_b_rows = df[df["b"] <= 0]

        raise ValueError(
            "判別楕円パラメータのbには，"
            "0より大きい値が必要です。\n\n"
            f"{invalid_b_rows.to_string(index=False)}"
        )

    # パラメータ表で最初に現れたグループ順を保持する。
    group_order = list(
        dict.fromkeys(
            df["Group"].tolist()
        )
    )

    df["Group"] = pd.Categorical(
        df["Group"],
        categories=group_order,
        ordered=True,
    )

    df = df.sort_values(
        [
            "Group",
            "Diagram",
        ],
        kind="stable",
    ).reset_index(drop=True)

    df["Group"] = df["Group"].astype(str)

    expected_diagrams = set(
        range(1, 11)
    )

    for group, group_df in df.groupby(
        "Group",
        sort=False,
    ):
        actual_diagrams = set(
            group_df["Diagram"].tolist()
        )

        if actual_diagrams != expected_diagrams:
            missing_diagrams = sorted(
                expected_diagrams
                - actual_diagrams
            )

            extra_diagrams = sorted(
                actual_diagrams
                - expected_diagrams
            )

            raise ValueError(
                f"{group}のDiagramがD1～D10で"
                "そろっていません。\n"
                f"不足：{missing_diagrams}\n"
                f"余分：{extra_diagrams}\n"
                f"現在：{sorted(actual_diagrams)}"
            )

        duplicated_mask = (
            group_df["Diagram"].duplicated(
                keep=False
            )
        )

        if duplicated_mask.any():
            duplicated_diagrams = sorted(
                group_df.loc[
                    duplicated_mask,
                    "Diagram",
                ]
                .astype(int)
                .unique()
                .tolist()
            )

            raise ValueError(
                f"{group}でDiagram番号が"
                "重複しています："
                f"{duplicated_diagrams}"
            )

        # D1～D10のX・Yの組合せを確認する。
        for _, parameter in group_df.iterrows():
            diagram = int(
                parameter["Diagram"]
            )

            actual_pair = (
                normalize_parameter_name(
                    parameter["X"]
                ),
                normalize_parameter_name(
                    parameter["Y"]
                ),
            )

            expected_pair = (
                EXPECTED_DIAGRAM_VARIABLES[
                    diagram
                ]
            )

            if actual_pair != expected_pair:
                raise ValueError(
                    f"{group}のD{diagram}の"
                    "X・Yが所定の組合せと異なります。\n"
                    f"現在：{actual_pair[0]} vs "
                    f"{actual_pair[1]}\n"
                    f"所定：{expected_pair[0]} vs "
                    f"{expected_pair[1]}"
                )

        actual_required_diagrams = set(
            group_df.loc[
                group_df["Required"] == 1,
                "Diagram",
            ]
            .astype(int)
            .tolist()
        )

        if (
            actual_required_diagrams
            != EXPECTED_REQUIRED_DIAGRAMS
        ):
            raise ValueError(
                f"{group}の必須Diagramが"
                "D1～D3になっていません。\n"
                f"現在："
                f"{sorted(actual_required_diagrams)}\n"
                "所定：[1, 2, 3]\n\n"
                "Mandatory／Required列は，"
                "D1～D3を1，D4～D10を0にしてください。"
            )

    return df


# ============================================================
# 5．定量値データの整形
# ============================================================

def prepare_quantitative_data(
    raw_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    定量値シートを判定用に整形する。

    判定用に次の5種を計算する。
    ・ln(100*K/Ca)
    ・Rb*
    ・Sr*
    ・Y*
    ・Zr*

    戻り値
    ------
    judgment_df：
        判定計算用データ

    original_df：
        元の定量値シート
    """

    original_df = raw_df.copy()

    column_map = {
        "Sample": find_column(
            raw_df,
            [
                "Sample",
                "sample",
                "試料",
                "試料名",
            ],
        ),
        "測定モード": find_column(
            raw_df,
            [
                "測定モード",
                "Method",
                "method",
                "Condition",
                "条件",
            ],
        ),
        "試料測定日時": find_column(
            raw_df,
            [
                "試料測定日時",
                "測定日時",
                "DateTime",
                "Datetime",
                "Date",
            ],
        ),
        "K": find_column(
            raw_df,
            [
                "K_ppm",
                "K",
            ],
        ),
        "Ca": find_column(
            raw_df,
            [
                "Ca_ppm",
                "Ca",
            ],
        ),
        "Mn": find_column(
            raw_df,
            [
                "Mn_ppm",
                "Mn",
            ],
        ),
        "Fe": find_column(
            raw_df,
            [
                "Fe_ppm",
                "Fe",
            ],
        ),
        "Zn": find_column(
            raw_df,
            [
                "Zn_ppm",
                "Zn",
            ],
        ),
        "Rb": find_column(
            raw_df,
            [
                "Rb_ppm",
                "Rb",
            ],
        ),
        "Sr": find_column(
            raw_df,
            [
                "Sr_ppm",
                "Sr",
            ],
        ),
        "Y": find_column(
            raw_df,
            [
                "Y_ppm",
                "Y",
            ],
        ),
        "Zr": find_column(
            raw_df,
            [
                "Zr_ppm",
                "Zr",
            ],
        ),
        "Nb": find_column(
            raw_df,
            [
                "Nb_ppm",
                "Nb",
            ],
        ),
    }

    rename_map = {
        original_name: standard_name
        for standard_name, original_name
        in column_map.items()
    }

    df = raw_df.rename(
        columns=rename_map
    ).copy()

    base_columns = [
        "Sample",
        "測定モード",
        "試料測定日時",
        "K",
        "Ca",
        "Mn",
        "Fe",
        "Zn",
        "Rb",
        "Sr",
        "Y",
        "Zr",
        "Nb",
    ]

    df = df[base_columns].copy()

    df["Sample"] = df["Sample"].map(
        clean_text
    )

    # Sampleが空欄の行は判定対象外とする。
    df = df[
        df["Sample"] != ""
    ].reset_index(drop=True)

    element_columns = [
        "K",
        "Ca",
        "Mn",
        "Fe",
        "Zn",
        "Rb",
        "Sr",
        "Y",
        "Zr",
        "Nb",
    ]

    for column in element_columns:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce",
        )

    denominator = (
        df["Rb"]
        + df["Sr"]
        + df["Y"]
        + df["Zr"]
    )

    with np.errstate(
        divide="ignore",
        invalid="ignore",
    ):
        df["Rb*"] = (
            100.0
            * df["Rb"]
            / denominator
        )

        df["Sr*"] = (
            100.0
            * df["Sr"]
            / denominator
        )

        df["Y*"] = (
            100.0
            * df["Y"]
            / denominator
        )

        df["Zr*"] = (
            100.0
            * df["Zr"]
            / denominator
        )

        df["ln(100*K/Ca)"] = np.log(
            100.0
            * df["K"]
            / df["Ca"]
        )

    df.replace(
        [
            np.inf,
            -np.inf,
        ],
        np.nan,
        inplace=True,
    )

    # 同一Sample内で1，2，3…と番号を付ける。
    # Excelの
    # =COUNTIF($O$45:O45,O45)
    # と同じ考え方。
    df["No."] = (
        df.groupby(
            "Sample",
            sort=False,
        )
        .cumcount()
        .add(1)
    )

    return df, original_df


# ============================================================
# 6．回転楕円の判定値
# ============================================================

def ellipse_value(
    x: float,
    y: float,
    x0: float,
    y0: float,
    a: float,
    b: float,
    theta_degree: float,
) -> float:
    """
    回転楕円の判定値を計算する。

    判定値 <= 1：楕円内
    判定値 > 1 ：楕円外
    """

    values = [
        x,
        y,
        x0,
        y0,
        a,
        b,
        theta_degree,
    ]

    if any(
        pd.isna(value)
        for value in values
    ):
        return np.nan

    theta = math.radians(
        theta_degree
    )

    cos_theta = math.cos(theta)
    sin_theta = math.sin(theta)

    dx = x - x0
    dy = y - y0

    rotated_x = (
        dx * cos_theta
        + dy * sin_theta
    )

    rotated_y = (
        -dx * sin_theta
        + dy * cos_theta
    )

    return math.sqrt(
        (rotated_x ** 2) / (a ** 2)
        +
        (rotated_y ** 2) / (b ** 2)
    )


# ============================================================
# 7．1組の定量分析値を判定
# ============================================================

def judge_one_analysis(
    row: pd.Series,
    grouped_parameters: list[
        tuple[str, pd.DataFrame]
    ],
) -> dict:
    """
    1組の定量分析値について，
    全化学組成グループを判定する。
    """

    all_group_results = []

    for group, group_df in grouped_parameters:
        diagram_results: dict[int, bool] = {}
        diagram_values: dict[int, float] = {}

        for _, parameter in group_df.iterrows():
            diagram = int(
                parameter["Diagram"]
            )

            x_name = normalize_parameter_name(
                parameter["X"]
            )

            y_name = normalize_parameter_name(
                parameter["Y"]
            )

            if x_name not in row.index:
                raise KeyError(
                    f"判別変数「{x_name}」を"
                    "計算できません。"
                )

            if y_name not in row.index:
                raise KeyError(
                    f"判別変数「{y_name}」を"
                    "計算できません。"
                )

            value = ellipse_value(
                x=row[x_name],
                y=row[y_name],
                x0=parameter["x0"],
                y0=parameter["y0"],
                a=parameter["a"],
                b=parameter["b"],
                theta_degree=parameter[
                    "theta"
                ],
            )

            diagram_values[diagram] = value

            diagram_results[diagram] = (
                False
                if pd.isna(value)
                else value <= ELLIPSE_LIMIT
            )

        required_diagrams = (
            group_df.loc[
                group_df["Required"] == 1,
                "Diagram",
            ]
            .astype(int)
            .tolist()
        )

        # D1～D3がすべて楕円内の場合のみ有効候補。
        candidate_is_valid = all(
            diagram_results[diagram]
            for diagram in required_diagrams
        )

        # D1～D10の楕円内の総数を得点とする。
        score = sum(
            1
            for is_inside
            in diagram_results.values()
            if is_inside
        )

        all_group_results.append(
            {
                "Group": group,
                "候補成立": candidate_is_valid,
                "得点": score,
                "Diagram判定": diagram_results,
                "Diagram判定値": diagram_values,
            }
        )

    valid_candidates = [
        result
        for result in all_group_results
        if result["候補成立"]
    ]

    if not valid_candidates:
        return {
            "判定結果": "",
            "得点": "",
            "有効/無効": "無効",
            "Diagram判定": {},
            "Diagram判定値": {},
            "全グループ結果": all_group_results,
        }

    # 同点の場合は，
    # パラメータ表で先に並ぶグループを採用する。
    winner = max(
        valid_candidates,
        key=lambda result: result["得点"],
    )

    return {
        "判定結果": winner["Group"],
        "得点": winner["得点"],
        "有効/無効": "有効",
        "Diagram判定": winner[
            "Diagram判定"
        ],
        "Diagram判定値": winner[
            "Diagram判定値"
        ],
        "全グループ結果": all_group_results,
    }


# ============================================================
# 8．Excel書式
# ============================================================

def autosize_worksheet(
    worksheet,
    maximum_width: int = 35,
) -> None:
    """
    列幅を内容に合わせて調整する。
    """

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            max_length = max(
                max_length,
                len(str(cell.value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                max_length + 2,
                8,
            ),
            maximum_width,
        )


def style_output_workbook(
    output_path: str,
) -> None:
    """
    出力Excelの書式を整える。
    """

    workbook = load_workbook(
        output_path
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    for sheet_name in [
        "判定結果",
        "判定詳細",
        "定量値",
    ]:
        worksheet = workbook[
            sheet_name
        ]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = (
                center_alignment
            )

        autosize_worksheet(
            worksheet
        )

    # 判定詳細のD1～D10を中央揃え。
    detail_sheet = workbook[
        "判定詳細"
    ]

    for row in detail_sheet.iter_rows(
        min_row=2,
        min_col=5,
        max_col=14,
    ):
        for cell in row:
            cell.alignment = (
                center_alignment
            )

    workbook.save(
        output_path
    )


# ============================================================
# Streamlit／Webアプリから呼び出す処理
# ============================================================

def create_group_judgment(
    parameter_content: bytes,
    quantitative_content: bytes,
    quantitative_file_name: str,
) -> dict:
    """2つの入力Excelを判定し、出力Excelをbytesで返す。"""
    if not parameter_content:
        raise ValueError("判別楕円パラメーター.xlsxが空です。")
    if not quantitative_content:
        raise ValueError("日付_定量値.xlsxが空です。")

    output_file_name = make_output_file_name(
        quantitative_file_name
    )

    parameter_bytes = io.BytesIO(parameter_content)
    parameter_excel = pd.ExcelFile(parameter_bytes)
    parameter_sheet_name = parameter_excel.sheet_names[0]
    parameter_bytes.seek(0)
    raw_parameter_df = pd.read_excel(
        parameter_bytes,
        sheet_name=parameter_sheet_name,
    )
    parameter_df = prepare_parameters(raw_parameter_df)
    grouped_parameters = list(
        parameter_df.groupby(
            "Group",
            sort=False,
            observed=True,
        )
    )
    group_names = [
        group
        for group, _ in grouped_parameters
    ]

    quantitative_bytes = io.BytesIO(quantitative_content)
    quantitative_excel = pd.ExcelFile(quantitative_bytes)
    if INPUT_SHEET_NAME not in quantitative_excel.sheet_names:
        raise ValueError(
            f"定量値ファイルに「{INPUT_SHEET_NAME}」シートがありません。\n"
            f"実際のシート：{quantitative_excel.sheet_names}"
        )

    quantitative_bytes.seek(0)
    raw_quantitative_df = pd.read_excel(
        quantitative_bytes,
        sheet_name=INPUT_SHEET_NAME,
    )
    judgment_df, original_quantitative_df = prepare_quantitative_data(
        raw_quantitative_df
    )
    if judgment_df.empty:
        raise ValueError(
            "「定量値」シートに判定対象となるデータがありません。"
        )

    result_rows = []
    detail_rows = []
    for _, row in judgment_df.iterrows():
        judgment = judge_one_analysis(
            row=row,
            grouped_parameters=grouped_parameters,
        )
        sample = row["Sample"]
        number = int(row["No."])

        result_rows.append(
            {
                "Sample": sample,
                "分析値No.": number,
                "判定結果": judgment["判定結果"],
                "得点": judgment["得点"],
                "有効/無効": judgment["有効/無効"],
            }
        )

        detail_row = {
            "Sample": sample,
            "分析値No.": number,
            "判定結果": judgment["判定結果"],
            "得点": judgment["得点"],
        }
        for diagram in range(1, 11):
            if judgment["有効/無効"] == "有効":
                inside = judgment["Diagram判定"].get(
                    diagram,
                    False,
                )
                detail_row[f"D{diagram}"] = (
                    INSIDE_MARK if inside else OUTSIDE_MARK
                )
            else:
                detail_row[f"D{diagram}"] = ""
        detail_rows.append(detail_row)

    result_columns = [
        "Sample",
        "分析値No.",
        "判定結果",
        "得点",
        "有効/無効",
    ]
    detail_columns = [
        "Sample",
        "分析値No.",
        "判定結果",
        "得点",
        *[f"D{diagram}" for diagram in range(1, 11)],
    ]
    result_df = pd.DataFrame(
        result_rows,
        columns=result_columns,
    )
    detail_df = pd.DataFrame(
        detail_rows,
        columns=detail_columns,
    )

    with tempfile.TemporaryDirectory() as temporary_directory:
        output_path = Path(temporary_directory) / output_file_name
        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
        ) as writer:
            result_df.to_excel(
                writer,
                sheet_name="判定結果",
                index=False,
            )
            detail_df.to_excel(
                writer,
                sheet_name="判定詳細",
                index=False,
            )
            original_quantitative_df.to_excel(
                writer,
                sheet_name="定量値",
                index=False,
            )

        style_output_workbook(str(output_path))
        output_content = output_path.read_bytes()

    valid_count = int(
        (result_df["有効/無効"] == "有効").sum()
    )
    invalid_count = int(
        (result_df["有効/無効"] == "無効").sum()
    )

    return {
        "file_name": output_file_name,
        "content": output_content,
        "analysis_count": len(result_df),
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "group_count": len(group_names),
        "group_names": group_names,
    }
