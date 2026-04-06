import csv
from io import StringIO, BytesIO

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="NEX DE 強度→定量値変換アプリ（長崎大ver1）",
    page_icon="🧪",
    layout="wide"
)

st.markdown("""
<div class="app-title">NEX DE 強度→定量値変換アプリ（長崎大ver1）</div>
<div class="app-subtitle">CSV extraction · drift correction · quantitative conversion · Excel export</div>
""", unsafe_allow_html=True)
st.write(
    "NEX DEから出力したCSVファイルをアップロードすると，"
    "強度データの抽出，ドリフト補正，定量値の算出を自動で行い，"
    "ボタンを押すことで結果をExcelファイルとしてダウンロードできます。"
)
st.write("注意）ドリフト補正のため「QC-2」のデータはCSVファイルに必ず含めてください。")

st.markdown("""
<style>
/* 全体の余白 */
.block-container {
    padding-top: 2rem;
    padding-bottom: 2rem;
}

/* チャットメッセージ */
[data-testid="stChatMessage"] {
    border-radius: 18px;
    padding: 0.8rem 1rem;
    margin-bottom: 0.8rem;
    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.06);
}

/* assistant側 */
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-assistant"]) {
    background: linear-gradient(135deg, #f7fbff 0%, #eef7ff 100%);
    border: 1px solid #d7e9f7;
}

/* user側 */
[data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]) {
    background: linear-gradient(135deg, #fcfcfc 0%, #f5f5f5 100%);
    border: 1px solid #e5e7eb;
}

/* アバター背景を消す */
[data-testid="stChatMessageAvatar"] {
    background: transparent !important;
    border: none !important;
}

/* データフレーム */
[data-testid="stDataFrame"] {
    border-radius: 14px;
    overflow: hidden;
}

/* ダウンロードボタン */
.stDownloadButton button {
    border-radius: 12px !important;
    border: 1px solid #cfd8e3 !important;
    background: linear-gradient(135deg, #ffffff 0%, #f4f8fc 100%) !important;
    color: #1f2937 !important;
    font-weight: 600 !important;
    padding: 0.6rem 1rem !important;
    box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06) !important;
}

.stDownloadButton button:hover {
    border-color: #93c5fd !important;
    color: #0f172a !important;
}

/* 入力欄 */
[data-testid="stChatInput"] {
    border-radius: 14px;
}

/* カスタムタイトル */
.app-title {
    font-size: 4.4rem;
    font-weight: 800;
    letter-spacing: -0.02em;
    line-height: 1.18;
    padding-top: 0.4rem;
    padding-bottom: 0.4rem;
    margin-top: 0.2rem;
    margin-bottom: 0.6rem;
    color: #2f3342;
}

.app-subtitle {
    font-size: 1rem;
    color: #7b8190;
    margin-bottom: 1.2rem;
}
</style>
""", unsafe_allow_html=True)

# =========================
# 1. QC-2 基準強度（cps/μA）
# =========================
reference_qc2 = {
    "Mid-Z_K-Kα": 4.27826,
    "Mid-Z_Ca-Kβ1": 0.15422,
    "Mid-Z_Mn-Kα": 0.90407,
    "Mid-Z_Fe-Kβ1": 3.15459,
    "High-Z_Zn-Kα": 0.01778,
    "High-Z_Rb-Kα": 0.22959,
    "High-Z_Sr-Kα": 0.07396,
    "High-Z_Y-Kα": 0.09792,
    "High-Z_Zr-Kα": 0.17312,
    "High-Z_Nb-Kα": 0.10719,
    "High-Z_Ag-Kα": 1.53429,
}
target_cols = list(reference_qc2.keys())

# =========================
# 2. 検量線定数
# ※ Zn も Ag コンプトン内標準化する前提
# =========================
B = {
    "K": 9312.59,
    "Ca": 30145.8,
    "Mn": 746.088,
    "Fe": 4061.51,
    "Zn": 7048.49,
    "Rb": 1360.26,
    "Sr": 1112.8,
    "Y": 898.725,
    "Zr": 810.974,
    "Nb": 737.066,
}

C = {
    "K": -1378.82,
    "Ca": -204.571,
    "Mn": -315.439,
    "Fe": -5513.26,
    "Zn": -44.4817,
    "Rb": -16.0436,
    "Sr": -11.9096,
    "Y": -12.6423,
    "Zr": -18.6844,
    "Nb": -31.1933,
}

# =========================
# 3. 重なり補正係数
# =========================
BG14 = -0.113036
BG15 = -0.121299
BG16 = -0.161034


def read_uploaded_csv(uploaded_file):
    raw = uploaded_file.getvalue()
    encodings = ["cp932", "utf-8-sig", "utf-8"]
    last_error = None

    for enc in encodings:
        try:
            text = raw.decode(enc, errors="replace")
            return list(csv.reader(StringIO(text)))
        except Exception as e:
            last_error = e

    raise ValueError(f"CSVを読み込めませんでした: {last_error}")


def clean_numeric(x):
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return x

    s = str(x).strip()
    if s == "":
        return None

    parts = s.split()
    if len(parts) >= 2:
        try:
            return float(parts[-1])
        except Exception:
            pass

    try:
        return float(s)
    except Exception:
        return None


def extract_cps_blocks(rows):
    cps_indices = []
    for i, row in enumerate(rows):
        normalized = [str(x).strip().lower() for x in row]
        if "cps/μa" in normalized:
            cps_indices.append(i)

    if not cps_indices:
        raise ValueError("cps/μA 行が見つかりません")

    all_data = []
    columns = None

    for idx in cps_indices:
        if idx < 2:
            continue

        header1 = rows[idx - 2]
        header2 = rows[idx - 1]

        if columns is None:
            new_cols = []
            seen = {}

            for h1, h2 in zip(header1[3:], header2[3:]):
                h1 = str(h1).strip()
                h2 = str(h2).strip()

                if h1 and h2:
                    name = f"{h1}_{h2}"
                elif h2:
                    name = h2
                else:
                    name = h1

                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 1

                new_cols.append(name)

            columns = ["Sample", "Method", "Date"] + new_cols

        for row in rows[idx + 1:]:
            if len(row) == 0 or all(str(x).strip() == "" for x in row):
                break

            row = list(row)

            if len(row) < len(columns):
                row = row + [""] * (len(columns) - len(row))
            elif len(row) > len(columns):
                row = row[:len(columns)]

            all_data.append(row)

    if not all_data:
        raise ValueError("cps/μA データが見つかりません")

    df = pd.DataFrame(all_data, columns=columns)

    numeric_df = df.iloc[:, 3:].copy()
    for col in numeric_df.columns:
        numeric_df[col] = numeric_df[col].map(clean_numeric)
        numeric_df[col] = pd.to_numeric(numeric_df[col], errors="coerce")

    df = pd.concat([df.iloc[:, :3].copy(), numeric_df], axis=1)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    return df


def calculate_drift_factors(df):
    sample_norm = (
        df["Sample"]
        .astype(str)
        .str.strip()
        .str.upper()
        .str.replace("-", "", regex=False)
        .str.replace(" ", "", regex=False)
    )

    qc2_rows = df[sample_norm == "QC2"]

    if qc2_rows.empty:
        raise ValueError("QC-2 または QC2 の行が見つかりません")

    qc2_measured = qc2_rows.iloc[0]

    drift_factors = {}
    for col in target_cols:
        if col not in df.columns:
            raise ValueError(f"必要な列が見つかりません: {col}")

        measured_val = qc2_measured[col]
        ref_val = reference_qc2[col]

        if pd.isna(measured_val) or measured_val == 0:
            drift_factors[col] = None
        else:
            drift_factors[col] = ref_val / measured_val

    return drift_factors


def apply_drift_and_quantification(df):
    drift_factors = calculate_drift_factors(df)

    df_corrected = df.copy()
    for col in target_cols:
        factor = drift_factors[col]
        if factor is not None:
            df_corrected[col] = df_corrected[col] * factor

    ag_col = "High-Z_Ag-Kα"
    if ag_col not in df_corrected.columns:
        raise ValueError(f"Ag 内標準列が見つかりません: {ag_col}")

    # K〜Fe
    df_corrected["K"] = B["K"] * df_corrected["Mid-Z_K-Kα"] + C["K"]
    df_corrected["Ca"] = B["Ca"] * df_corrected["Mid-Z_Ca-Kβ1"] + C["Ca"]
    df_corrected["Mn"] = B["Mn"] * df_corrected["Mid-Z_Mn-Kα"] + C["Mn"]
    df_corrected["Fe"] = B["Fe"] * df_corrected["Mid-Z_Fe-Kβ1"] + C["Fe"]

    # Zn も Ag コンプトン内標準化
    df_corrected["Zn"] = B["Zn"] * (df_corrected["High-Z_Zn-Kα"] / df_corrected[ag_col]) + C["Zn"]

    # Rb, Sr
    df_corrected["Rb"] = B["Rb"] * (df_corrected["High-Z_Rb-Kα"] / df_corrected[ag_col]) + C["Rb"]
    df_corrected["Sr"] = B["Sr"] * (df_corrected["High-Z_Sr-Kα"] / df_corrected[ag_col]) + C["Sr"]

    # Y, Zr, Nb
    df_corrected["Y"] = (
        B["Y"] * (df_corrected["High-Z_Y-Kα"] / df_corrected[ag_col])
        + C["Y"]
        + df_corrected["Rb"] * BG14
    )
    df_corrected["Zr"] = (
        B["Zr"] * (df_corrected["High-Z_Zr-Kα"] / df_corrected[ag_col])
        + C["Zr"]
        + df_corrected["Sr"] * BG15
    )
    df_corrected["Nb"] = (
        B["Nb"] * (df_corrected["High-Z_Nb-Kα"] / df_corrected[ag_col])
        + C["Nb"]
        + df_corrected["Y"] * BG16
    )

    result_cols = [
        "Sample", "Method", "Date",
        "K", "Ca", "Mn", "Fe", "Zn",
        "Rb", "Sr", "Y", "Zr", "Nb",
    ]

    df_result = df_corrected[result_cols].copy()

    df_result = df_result.rename(columns={
        "K": "K ppm",
        "Ca": "Ca ppm",
        "Mn": "Mn ppm",
        "Fe": "Fe ppm",
        "Zn": "Zn ppm",
        "Rb": "Rb ppm",
        "Sr": "Sr ppm",
        "Y": "Y ppm",
        "Zr": "Zr ppm",
        "Nb": "Nb ppm",
    })

    ppm_cols = [
        "K ppm", "Ca ppm", "Mn ppm", "Fe ppm", "Zn ppm",
        "Rb ppm", "Sr ppm", "Y ppm", "Zr ppm", "Nb ppm"
    ]

    df_result[ppm_cols] = df_result[ppm_cols].round(2)
    df_result["Date"] = pd.to_datetime(df_result["Date"], errors="coerce")

    # 古い → 新しい
    df_result = df_result.sort_values(
        "Date",
        ascending=True,
        na_position="last"
    ).reset_index(drop=True)

    return df_result, drift_factors


def make_display_df(df):
    df_display = df.copy()

    ppm_cols = [
        "K ppm", "Ca ppm", "Mn ppm", "Fe ppm", "Zn ppm",
        "Rb ppm", "Sr ppm", "Y ppm", "Zr ppm", "Nb ppm"
    ]

    for col in ppm_cols:
        if col in df_display.columns:
            df_display[col] = df_display[col].map(
                lambda x: "" if pd.isna(x) else f"{x:.2f}"
            )

    if "Date" in df_display.columns:
        df_display["Date"] = pd.to_datetime(
            df_display["Date"], errors="coerce"
        ).dt.strftime("%Y-%m-%d %H:%M:%S")

    return df_display


def to_excel_bytes(df):
    df_export = df.copy()

    ppm_cols = [
        "K ppm", "Ca ppm", "Mn ppm", "Fe ppm", "Zn ppm",
        "Rb ppm", "Sr ppm", "Y ppm", "Zr ppm", "Nb ppm"
    ]

    df_export["Date"] = pd.to_datetime(df_export["Date"], errors="coerce")

    # Excel 出力直前にも古い → 新しいで固定
    df_export = df_export.sort_values(
        "Date",
        ascending=True,
        na_position="last"
    ).reset_index(drop=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="quantified_results")

        ws = writer.sheets["quantified_results"]

        # Date列の表示形式
        if "Date" in df_export.columns:
            date_col_idx = df_export.columns.get_loc("Date") + 1
            for row in range(2, len(df_export) + 2):
                ws.cell(row=row, column=date_col_idx).number_format = "yyyy-mm-dd hh:mm:ss"

        # ppm列は常に小数点以下2桁表示
        for col_name in ppm_cols:
            if col_name in df_export.columns:
                col_idx = df_export.columns.get_loc(col_name) + 1
                for row in range(2, len(df_export) + 2):
                    ws.cell(row=row, column=col_idx).number_format = "0.00"

        # 列幅調整
        for i, col_name in enumerate(df_export.columns, start=1):
            if col_name == "Date":
                ws.column_dimensions[get_column_letter(i)].width = 22
            elif "ppm" in col_name:
                ws.column_dimensions[get_column_letter(i)].width = 12
            else:
                ws.column_dimensions[get_column_letter(i)].width = 18

    output.seek(0)
    return output.getvalue()


# =========================
# 4. 対話UI
# =========================
uploaded = st.file_uploader("NEX DEから出力したCSVファイルをアップロード", type=["csv"])

if "messages" not in st.session_state:
    st.session_state.messages = []

if "df_result" not in st.session_state:
    st.session_state.df_result = None

if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

if "drift_factors" not in st.session_state:
    st.session_state.drift_factors = None

for msg in st.session_state.messages:
    avatar = "🧑‍🔬" if msg["role"] == "user" else "🧪"
    with st.chat_message(msg["role"], avatar=avatar):
        st.write(msg["content"])

prompt = st.chat_input(
    "例：定量計算して / ドリフト補正係数を見せて / 検量線定数を見せて / 重なり補正係数を見せて"
)

if prompt:
    st.session_state.messages.append({"role": "user", "content": prompt})

    with st.chat_message("user", avatar="🧑‍🔬"):
        st.write(prompt)

    lower_prompt = prompt.lower().strip()

    if uploaded is None and (
        ("定量" in prompt) or
        ("計算" in prompt) or
        ("結果" in prompt) or
        ("表" in prompt) or
        ("補正係数" in prompt) or
        ("drift" in lower_prompt)
    ):
        reply = "まず，NEX DEから出力したCSVファイルをアップロードしてください。"
        with st.chat_message("assistant", avatar="⚠️"):
            st.write(reply)
        st.session_state.messages.append({"role": "assistant", "content": reply})

    elif ("定量" in prompt) or ("計算" in prompt):
        try:
            rows = read_uploaded_csv(uploaded)
            df_counts = extract_cps_blocks(rows)
            df_result, drift_factors = apply_drift_and_quantification(df_counts)
            excel_bytes = to_excel_bytes(df_result)

            st.session_state.df_result = df_result
            st.session_state.excel_bytes = excel_bytes
            st.session_state.drift_factors = drift_factors

            reply = (
                f"定量計算が完了しました。{len(df_result)} 行の結果を作成しました。"
                " 下に結果表を表示し，Excelもダウンロードできます。"
            )

            with st.chat_message("assistant", avatar="🧪"):
                st.write(reply)
                st.dataframe(make_display_df(df_result), use_container_width=True)

            st.session_state.messages.append({"role": "assistant", "content": reply})

        except Exception as e:
            reply = f"処理中にエラーが出ました: {e}"
            with st.chat_message("assistant", avatar="⚠️"):
                st.write(reply)
            st.session_state.messages.append({"role": "assistant", "content": reply})

    elif ("ドリフト補正係数" in prompt) or ("drift" in lower_prompt):
        if st.session_state.drift_factors is None:
            reply = "まだドリフト補正係数がありません。先に「定量計算して」と入力してください。"
            with st.chat_message("assistant", avatar="⚠️"):
                st.write(reply)
            st.session_state.messages.append({"role": "assistant", "content": reply})
        else:
            with st.chat_message("assistant", avatar="🧪"):
                st.write("ドリフト補正係数を表示します。")
                drift_df = pd.DataFrame({
                    "Line": list(st.session_state.drift_factors.keys()),
                    "Drift factor": list(st.session_state.drift_factors.values()),
                })
                st.dataframe(drift_df, use_container_width=True)

            st.session_state.messages.append(
                {"role": "assistant", "content": "ドリフト補正係数を表示しました。"}
            )

    elif ("検量線定数" in prompt) or ("bとc" in lower_prompt) or ("b c" in lower_prompt):
        with st.chat_message("assistant", avatar="🧪"):
            st.write("検量線定数（B, C）を表示します。")
            calib_df = pd.DataFrame({
                "Element": list(B.keys()),
                "B": [B[k] for k in B.keys()],
                "C": [C[k] for k in B.keys()],
            })
            st.dataframe(calib_df, use_container_width=True)

        st.session_state.messages.append(
            {"role": "assistant", "content": "検量線定数（B, C）を表示しました。"}
        )

    elif ("重なり補正係数" in prompt) or ("bg14" in lower_prompt) or ("bg15" in lower_prompt) or ("bg16" in lower_prompt):
        with st.chat_message("assistant", avatar="🧪"):
            st.write("重なり補正係数を表示します。")
            overlap_df = pd.DataFrame({
                "Coefficient": ["BG14", "BG15", "BG16"],
                "Value": [BG14, BG15, BG16],
                "Meaning": [
                    "Y 補正に使う Rb 項の係数",
                    "Zr 補正に使う Sr 項の係数",
                    "Nb 補正に使う Y 項の係数",
                ],
            })
            st.dataframe(overlap_df, use_container_width=True)

        st.session_state.messages.append(
            {"role": "assistant", "content": "重なり補正係数を表示しました。"}
        )

    elif ("qc-2基準強度" in lower_prompt) or ("qc2基準強度" in lower_prompt) or ("基準強度" in prompt):
        with st.chat_message("assistant", avatar="🧪"):
            st.write("QC-2 基準強度を表示します。")
            qc2_df = pd.DataFrame({
                "Line": list(reference_qc2.keys()),
                "Reference intensity (cps/μA)": list(reference_qc2.values()),
            })
            st.dataframe(qc2_df, use_container_width=True)

        st.session_state.messages.append(
            {"role": "assistant", "content": "QC-2 基準強度を表示しました。"}
        )

    elif ("結果" in prompt) or ("表" in prompt):
        if st.session_state.df_result is None:
            reply = "まだ結果がありません。先に「定量計算して」と入力してください。"
            with st.chat_message("assistant", avatar="⚠️"):
                st.write(reply)
            st.session_state.messages.append({"role": "assistant", "content": reply})
        else:
            with st.chat_message("assistant", avatar="🧪"):
                st.write("現在の定量結果を表示します。")
                st.dataframe(make_display_df(st.session_state.df_result), use_container_width=True)

            st.session_state.messages.append(
                {"role": "assistant", "content": "定量結果を表示しました。"}
            )

    else:
        reply = (
            "現在対応している指示は，"
            "「定量計算して」「ドリフト補正係数を見せて」「結果を見せて」"
            "「検量線定数を見せて」「重なり補正係数を見せて」"
            "「QC-2基準強度を見せて」です。"
        )
        with st.chat_message("assistant", avatar="🧪"):
            st.write(reply)
        st.session_state.messages.append({"role": "assistant", "content": reply})

if st.session_state.df_result is not None:
    if uploaded is not None:
        base_name = uploaded.name.rsplit(".", 1)[0]
    else:
        base_name = "result"

    st.download_button(
        label="Excelをダウンロード",
        data=st.session_state.excel_bytes,
        file_name=f"{base_name}_quantified_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
