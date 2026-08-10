"""
Google Colab用：EDX CSVから定量値Excelまでを一括作成する。

処理:
1. NEX DEから出力した日付.csvをアップロード
2. 機関ごとの定量計算パラメーター.xlsxをアップロード
3. パラメーターに指定されたQC測定からドリフト補正係数を算出
4. 指定された元素にAg内標準補正を適用
5. 検量線定数B・Cから定量値を算出
6. 指定された重なり補正を適用

Colabでは、このコード全体を1つのセルに貼り付けて実行する。
CSVとパラメーターExcelを順番にアップロードすると、
日付_定量値.xlsxが自動的にダウンロードされる。
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ELEMENTS = ("K", "Ca", "Mn", "Fe", "Zn", "Rb", "Sr", "Y", "Zr", "Nb", "Ag")
QUANT_ELEMENTS = ELEMENTS[:-1]
DIRECT_ELEMENTS = {"K", "Ca", "Mn", "Fe"}
AG_NORMALIZED_ELEMENTS = {"Zn", "Rb", "Sr", "Y", "Zr", "Nb"}
QC_MODE = "obart_prec3_air"
QC_SAMPLE_NAME = "QC-2"

REFERENCE_INTENSITY = {
    "K": 4.27826,
    "Ca": 0.15422,
    "Mn": 0.90407,
    "Fe": 3.15459,
    "Zn": 0.01778,
    "Rb": 0.22959,
    "Sr": 0.07396,
    "Y": 0.09792,
    "Zr": 0.17312,
    "Nb": 0.10719,
    "Ag": 1.53429,
}

CALIBRATION_B = {
    "K": 9312.59,
    "Ca": 30145.8,
    "Mn": 746.088,
    "Fe": 4061.51,
    "Zn": 7048.49,
    "Rb": 1360.26,
    "Sr": 1112.8,
    "Y": 898.725,
    "Zr": 810.974,
    "Nb": 737.066,
}

CALIBRATION_C = {
    "K": -1378.82,
    "Ca": -204.571,
    "Mn": -315.439,
    "Fe": -5513.26,
    "Zn": -44.4817,
    "Rb": -16.0436,
    "Sr": -11.9096,
    "Y": -12.6423,
    "Zr": -18.6844,
    "Nb": -31.1933,
}

# 対象元素: (重なり元素, 重なり補正係数)
OVERLAP_CORRECTION = {
    "Y": ("Rb", -0.113036),
    "Zr": ("Sr", -0.121299),
    "Nb": ("Y", -0.161034),
}

_CONVERSION_LOCK = Lock()


def normalize_parameter_text(value) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    return re.sub(r"[\s_・=()（）]+", "", text)


def make_output_name(input_name: str) -> str:
    input_stem = Path(input_name).stem
    base_name = re.sub(r"\s*\(\d+\)$", "", input_stem).strip()
    return f"{base_name}_定量値.xlsx"


def read_numeric(value, description: str) -> float:
    if value is None or str(value).strip() == "":
        raise ValueError(f"{description}が空欄です。")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{description}を数値として読めません: {value!r}") from exc
    return number


def find_parameter_value(sheet, labels: tuple[str, ...], required=True):
    targets = {normalize_parameter_text(label) for label in labels}
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_parameter_text(cell.value) in targets:
                return sheet.cell(cell.row, cell.column + 1).value
    if required:
        raise ValueError(f"パラメーターファイルに項目がありません: {labels}")
    return ""


def load_calculation_parameters(parameter_bytes: bytes) -> dict:
    workbook = load_workbook(io.BytesIO(parameter_bytes), data_only=True)
    if "定量計算パラメーター" not in workbook.sheetnames:
        raise ValueError(
            "パラメーターファイルに「定量計算パラメーター」シートがありません。\n"
            f"実際のシート: {workbook.sheetnames}"
        )
    sheet = workbook["定量計算パラメーター"]

    institution = str(
        find_parameter_value(sheet, ("機関名",), required=False) or ""
    ).strip()
    version = str(
        find_parameter_value(
            sheet,
            ("パラメーターバージョン",),
            required=False,
        ) or ""
    ).strip()
    qc_sample_name = str(
        find_parameter_value(
            sheet,
            ("ドリフト補正資料", "QC試料名"),
        )
    ).strip()
    qc_mode = str(
        find_parameter_value(sheet, ("測定モード", "QC測定モード"))
    ).strip()
    if not qc_sample_name:
        raise ValueError("ドリフト補正資料（QC試料名）が空欄です。")
    if not qc_mode:
        raise ValueError("測定モードが空欄です。")

    expected_headers = {
        "元素": "element",
        "QC-2基準強度": "reference",
        "検量線定数B": "b",
        "検量線定数C": "c",
        "補正方法": "method",
        "重なり元素": "overlap_element",
        "重なり補正係数": "overlap_coefficient",
    }
    normalized_headers = {
        normalize_parameter_text(header): key
        for header, key in expected_headers.items()
    }

    header_row = None
    column_map = {}
    for row in sheet.iter_rows():
        row_map = {}
        for cell in row:
            normalized = normalize_parameter_text(cell.value)
            if normalized in normalized_headers:
                row_map[normalized_headers[normalized]] = cell.column
        if {"element", "reference", "b", "c", "method"}.issubset(row_map):
            header_row = row[0].row
            column_map = row_map
            break
    if header_row is None:
        raise ValueError(
            "元素パラメーター表の見出しを確認できません。"
            "必要列: 元素、QC-2基準強度、検量線定数B、"
            "検量線定数C、補正方法、重なり補正係数"
        )

    reference_intensity = {}
    calibration_b = {}
    calibration_c = {}
    direct_elements = set()
    ag_normalized_elements = set()
    overlap_correction = {}

    for row_number in range(header_row + 1, sheet.max_row + 1):
        element_value = sheet.cell(
            row_number,
            column_map["element"],
        ).value
        element = str(element_value or "").strip()
        if not element:
            continue
        if element not in ELEMENTS:
            raise ValueError(f"未対応の元素があります: {element}")
        if element in reference_intensity:
            raise ValueError(f"元素が重複しています: {element}")

        reference = read_numeric(
            sheet.cell(row_number, column_map["reference"]).value,
            f"{element}のQC基準強度",
        )
        if reference <= 0:
            raise ValueError(f"{element}のQC基準強度は0より大きくしてください。")
        reference_intensity[element] = reference

        if element != "Ag":
            calibration_b[element] = read_numeric(
                sheet.cell(row_number, column_map["b"]).value,
                f"{element}の検量線定数B",
            )
            calibration_c[element] = read_numeric(
                sheet.cell(row_number, column_map["c"]).value,
                f"{element}の検量線定数C",
            )

        method = unicodedata.normalize(
            "NFKC",
            str(sheet.cell(row_number, column_map["method"]).value or ""),
        ).replace(" ", "").replace("　", "")
        if "ドリフト補正" not in method:
            raise ValueError(
                f"{element}の補正方法に「ドリフト補正」がありません: {method!r}"
            )
        if element != "Ag":
            if "Ag内標準補正" in method:
                ag_normalized_elements.add(element)
            else:
                direct_elements.add(element)

        coefficient_column = column_map.get("overlap_coefficient")
        coefficient_value = (
            sheet.cell(row_number, coefficient_column).value
            if coefficient_column
            else None
        )
        overlap_column = column_map.get("overlap_element")
        overlap_element = (
            str(sheet.cell(row_number, overlap_column).value or "").strip()
            if overlap_column
            else ""
        )
        if not overlap_element:
            overlap_match = re.search(r"([A-Z][a-z]?)重なり補正", method)
            if overlap_match:
                overlap_element = overlap_match.group(1)

        if coefficient_value not in (None, ""):
            coefficient = read_numeric(
                coefficient_value,
                f"{element}の重なり補正係数",
            )
            if not overlap_element:
                raise ValueError(
                    f"{element}には重なり補正係数がありますが、"
                    "重なり元素を確認できません。"
                )
            if overlap_element not in QUANT_ELEMENTS:
                raise ValueError(
                    f"{element}の重なり元素が未対応です: {overlap_element}"
                )
            overlap_correction[element] = (overlap_element, coefficient)
        elif overlap_element:
            raise ValueError(
                f"{element}には重なり元素がありますが、補正係数が空欄です。"
            )

    missing_elements = [
        element for element in ELEMENTS
        if element not in reference_intensity
    ]
    if missing_elements:
        raise ValueError(
            "パラメーター表に不足している元素があります: "
            + ", ".join(missing_elements)
        )
    if direct_elements | ag_normalized_elements != set(QUANT_ELEMENTS):
        raise ValueError("定量対象元素の補正方法をすべて指定してください。")

    return {
        "institution": institution,
        "version": version,
        "qc_sample_name": qc_sample_name,
        "qc_mode": qc_mode,
        "reference_intensity": reference_intensity,
        "calibration_b": calibration_b,
        "calibration_c": calibration_c,
        "direct_elements": direct_elements,
        "ag_normalized_elements": ag_normalized_elements,
        "overlap_correction": overlap_correction,
    }


def apply_calculation_parameters(parameters: dict) -> None:
    global QC_SAMPLE_NAME, QC_MODE
    global REFERENCE_INTENSITY, CALIBRATION_B, CALIBRATION_C
    global DIRECT_ELEMENTS, AG_NORMALIZED_ELEMENTS, OVERLAP_CORRECTION

    QC_SAMPLE_NAME = parameters["qc_sample_name"]
    QC_MODE = parameters["qc_mode"]
    REFERENCE_INTENSITY = parameters["reference_intensity"]
    CALIBRATION_B = parameters["calibration_b"]
    CALIBRATION_C = parameters["calibration_c"]
    DIRECT_ELEMENTS = parameters["direct_elements"]
    AG_NORMALIZED_ELEMENTS = parameters["ag_normalized_elements"]
    OVERLAP_CORRECTION = parameters["overlap_correction"]


def decode_csv(csv_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "cp932"):
        try:
            return csv_bytes.decode(encoding)
        except UnicodeDecodeError:
            pass
    raise ValueError("CSVの文字コードを判定できませんでした。")


def parse_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = str(value or "").strip()
    for date_format in (
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    raise ValueError(f"測定日時を読めません: {value!r}")


def normalize_qc_name(value) -> str:
    text = str(value or "").strip().upper()
    for character in ("-", "－", "−", "‐", "‑", "–", "—", " ", "　"):
        text = text.replace(character, "")
    return text


def is_qc(measurement: dict) -> bool:
    return (
        normalize_qc_name(measurement["sample"])
        == normalize_qc_name(QC_SAMPLE_NAME)
        and measurement["mode"].strip().lower() == QC_MODE.lower()
    )


def extract_measurements(csv_bytes: bytes) -> list[dict]:
    rows = list(csv.reader(io.StringIO(decode_csv(csv_bytes))))
    measurements = []
    in_intensity_block = False

    for row_number, source_row in enumerate(rows, start=1):
        row = source_row + [""] * max(0, 14 - len(source_row))

        # μが文字化けしても「cps/」で強度ブロックを識別する。
        if sum("cps/" in cell.strip().lower() for cell in row) >= 5:
            in_intensity_block = True
            continue

        if not any(cell.strip() for cell in row):
            in_intensity_block = False
            continue

        if not in_intensity_block:
            continue

        if not all(row[column].strip() for column in range(3)):
            continue

        intensities = {}
        for column, element in enumerate(ELEMENTS, start=3):
            text = row[column].strip()
            if not text:
                raise ValueError(
                    f"CSV {row_number}行目の{element}強度が空欄です。"
                )
            try:
                intensities[element] = float(text)
            except ValueError as exc:
                raise ValueError(
                    f"CSV {row_number}行目の{element}強度を数値として読めません。"
                ) from exc

        measurements.append(
            {
                "sample": row[0].strip(),
                "mode": row[1].strip(),
                "measured_at": parse_datetime(row[2]),
                "intensities": intensities,
            }
        )

    if not measurements:
        raise ValueError("CSV内に強度測定データが見つかりません。")
    return measurements


def calculate_quantitative_values(corrected: dict[str, float]) -> dict[str, float]:
    values = {}

    for element in ("K", "Ca", "Mn", "Fe", "Zn", "Rb", "Sr"):
        values[element] = (
            CALIBRATION_B[element] * corrected[element]
            + CALIBRATION_C[element]
        )

    overlap_element, coefficient = OVERLAP_CORRECTION["Y"]
    values["Y"] = (
        CALIBRATION_B["Y"] * corrected["Y"]
        + CALIBRATION_C["Y"]
        + values[overlap_element] * coefficient
    )

    overlap_element, coefficient = OVERLAP_CORRECTION["Zr"]
    values["Zr"] = (
        CALIBRATION_B["Zr"] * corrected["Zr"]
        + CALIBRATION_C["Zr"]
        + values[overlap_element] * coefficient
    )

    overlap_element, coefficient = OVERLAP_CORRECTION["Nb"]
    values["Nb"] = (
        CALIBRATION_B["Nb"] * corrected["Nb"]
        + CALIBRATION_C["Nb"]
        + values[overlap_element] * coefficient
    )
    return values


def calculate_all(measurements: list[dict]):
    qc_by_date = defaultdict(list)
    for measurement in measurements:
        if is_qc(measurement):
            qc_by_date[measurement["measured_at"].date()].append(measurement)

    for qc_list in qc_by_date.values():
        qc_list.sort(key=lambda item: item["measured_at"])

    if not qc_by_date:
        raise ValueError(
            f"QC試料「{QC_SAMPLE_NAME}」かつ測定モード「{QC_MODE}」の"
            "QC測定が見つかりません。"
        )

    coefficient_rows = []
    coefficients_by_date = defaultdict(list)

    for measurement_date in sorted(qc_by_date):
        for qc_number, qc in enumerate(qc_by_date[measurement_date], start=1):
            coefficients = {}
            for element in ELEMENTS:
                qc_intensity = qc["intensities"][element]
                if qc_intensity == 0:
                    raise ValueError(
                        f"{qc['measured_at']}の{element} QC強度が0です。"
                    )
                coefficients[element] = (
                    REFERENCE_INTENSITY[element] / qc_intensity
                )

            coefficients_by_date[measurement_date].append(
                (qc_number, qc, coefficients)
            )
            coefficient_rows.append(
                {
                    "date": measurement_date,
                    "qc_number": qc_number,
                    "qc": qc,
                    "coefficients": coefficients,
                }
            )

    corrected_rows = []
    missing_qc_dates = set()

    for sample in measurements:
        if is_qc(sample):
            continue

        measurement_date = sample["measured_at"].date()
        daily_coefficients = coefficients_by_date.get(measurement_date, [])
        if not daily_coefficients:
            missing_qc_dates.add(measurement_date)
            continue

        for qc_number, qc, coefficients in daily_coefficients:
            drift_corrected = {
                element: sample["intensities"][element] * coefficients[element]
                for element in ELEMENTS
            }

            corrected_ag = drift_corrected["Ag"]
            if corrected_ag == 0:
                raise ValueError(
                    f"{sample['sample']} / QC{qc_number}: "
                    "ドリフト補正後Ag強度が0です。"
                )

            corrected = {}
            for element in ELEMENTS:
                if element in DIRECT_ELEMENTS or element == "Ag":
                    corrected[element] = drift_corrected[element]
                elif element in AG_NORMALIZED_ELEMENTS:
                    corrected[element] = drift_corrected[element] / corrected_ag

            corrected_rows.append(
                {
                    "sample": sample,
                    "date": measurement_date,
                    "qc_number": qc_number,
                    "qc": qc,
                    "corrected": corrected,
                    "quantitative": calculate_quantitative_values(corrected),
                }
            )

    if not corrected_rows:
        raise ValueError("同日のQCを適用できる試料がありません。")

    return (
        coefficient_rows,
        corrected_rows,
        sorted(missing_qc_dates),
    )


def style_sheet(sheet, datetime_columns=(), date_columns=()):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="游ゴシック", size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font

    for column in datetime_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd h:mm"
    for column in date_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 11), 22
        )


def create_workbook(coefficient_rows, corrected_rows, missing_dates) -> Workbook:
    workbook = Workbook()

    qc_sheet = workbook.active
    qc_sheet.title = "QC補正係数"
    qc_sheet.append(
        [
            "測定日",
            "QC番号",
            "QC sample",
            "QC測定モード",
            "QC測定日時",
            *[f"{element}_補正係数" for element in ELEMENTS],
        ]
    )
    for row in coefficient_rows:
        qc_sheet.append(
            [
                row["date"],
                row["qc_number"],
                row["qc"]["sample"],
                row["qc"]["mode"],
                row["qc"]["measured_at"],
                *[row["coefficients"][element] for element in ELEMENTS],
            ]
        )
    style_sheet(qc_sheet, datetime_columns=("E",), date_columns=("A",))
    for row in qc_sheet.iter_rows(min_row=2, min_col=6, max_col=16):
        for cell in row:
            cell.number_format = "0.000000"

    corrected_sheet = workbook.create_sheet("補正後強度")
    corrected_sheet.append(
        [
            "sample",
            "測定モード",
            "試料測定日時",
            "測定日",
            "QC番号",
            "QC測定日時",
            "K_ドリフト補正後",
            "Ca_ドリフト補正後",
            "Mn_ドリフト補正後",
            "Fe_ドリフト補正後",
            "Zn_Ag内標準後",
            "Rb_Ag内標準後",
            "Sr_Ag内標準後",
            "Y_Ag内標準後",
            "Zr_Ag内標準後",
            "Nb_Ag内標準後",
            "Ag_ドリフト補正後",
        ]
    )
    for row in corrected_rows:
        corrected_sheet.append(
            [
                row["sample"]["sample"],
                row["sample"]["mode"],
                row["sample"]["measured_at"],
                row["date"],
                row["qc_number"],
                row["qc"]["measured_at"],
                *[row["corrected"][element] for element in ELEMENTS],
            ]
        )
    style_sheet(
        corrected_sheet,
        datetime_columns=("C", "F"),
        date_columns=("D",),
    )
    for row in corrected_sheet.iter_rows(min_row=2, min_col=7, max_col=17):
        for cell in row:
            cell.number_format = "0.000000"

    quantitative_sheet = workbook.create_sheet("定量値")
    quantitative_sheet.append(
        [
            "sample",
            "測定モード",
            "試料測定日時",
            *[f"{element}_ppm" for element in QUANT_ELEMENTS],
        ]
    )
    for row in corrected_rows:
        quantitative_sheet.append(
            [
                row["sample"]["sample"],
                row["sample"]["mode"],
                row["sample"]["measured_at"],
                *[row["quantitative"][element] for element in QUANT_ELEMENTS],
            ]
        )
    style_sheet(
        quantitative_sheet,
        datetime_columns=("C",),
    )
    for row in quantitative_sheet.iter_rows(min_row=2, min_col=4, max_col=13):
        for cell in row:
            cell.number_format = "0.000"

    constants_sheet = workbook.create_sheet("計算定数")
    constants_sheet.append(
        [
            "元素",
            "検量線定数B",
            "検量線定数C",
            "重なり元素",
            "重なり補正係数",
        ]
    )
    for element in QUANT_ELEMENTS:
        overlap_element, overlap_coefficient = OVERLAP_CORRECTION.get(
            element, ("", "")
        )
        constants_sheet.append(
            [
                element,
                CALIBRATION_B[element],
                CALIBRATION_C[element],
                overlap_element,
                overlap_coefficient,
            ]
        )
    style_sheet(constants_sheet)
    for row in constants_sheet.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"

    if missing_dates:
        warning_sheet = workbook.create_sheet("警告")
        warning_sheet.append(["同日のQCがなく、出力から除外した測定日"])
        for missing_date in missing_dates:
            warning_sheet.append([missing_date])
        style_sheet(warning_sheet, date_columns=("A",))

    # 「定量値」シートは、見出し・文字・数字をすべて中央揃えにする。
    for row in quantitative_sheet.iter_rows(
        min_row=1,
        max_row=quantitative_sheet.max_row,
        min_col=1,
        max_col=quantitative_sheet.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    # 出力には不要な中間確認用シートを除外する。
    workbook.remove(corrected_sheet)
    workbook.remove(constants_sheet)

    # 「定量値」シートを一番左へ移動し、Excelを開いたときにも表示する。
    workbook.move_sheet(
        quantitative_sheet,
        offset=-workbook.index(quantitative_sheet),
    )
    workbook.active = workbook.index(quantitative_sheet)

    return workbook


def convert_csv(csv_bytes: bytes) -> bytes:
    """将来Webアプリでも再利用できる、bytes入出力の一括変換関数。"""
    measurements = extract_measurements(csv_bytes)
    coefficient_rows, corrected_rows, missing_dates = calculate_all(measurements)
    workbook = create_workbook(
        coefficient_rows,
        corrected_rows,
        missing_dates,
    )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_quantitative_excel(
    parameter_content: bytes,
    csv_content: bytes,
    csv_file_name: str,
) -> dict:
    """パラメーターとCSVから、定量値Excelをbytesで作成する。"""
    if not parameter_content:
        raise ValueError("定量計算パラメーター.xlsxが空です。")
    if not csv_content:
        raise ValueError("日付.csvが空です。")

    parameters = load_calculation_parameters(parameter_content)

    # 計算条件は既存計算関数のグローバル定数へ適用するため、
    # 複数リクエストが同時に混ざらないよう一括処理をロックする。
    with _CONVERSION_LOCK:
        apply_calculation_parameters(parameters)
        measurements = extract_measurements(csv_content)
        coefficient_rows, corrected_rows, missing_dates = calculate_all(
            measurements
        )
        workbook = create_workbook(
            coefficient_rows,
            corrected_rows,
            missing_dates,
        )
        output = io.BytesIO()
        workbook.save(output)

    return {
        "file_name": make_output_name(csv_file_name),
        "content": output.getvalue(),
        "institution": parameters["institution"],
        "version": parameters["version"],
        "qc_sample_name": parameters["qc_sample_name"],
        "qc_mode": parameters["qc_mode"],
        "output_row_count": len(corrected_rows),
        "missing_dates": missing_dates,
    }
